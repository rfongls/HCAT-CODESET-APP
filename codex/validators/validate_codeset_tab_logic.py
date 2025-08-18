from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from codeset_ui_app.utils.xlsx_sanitizer import strip_invalid_font_families

DEFAULT_DEFINITION = Path(__file__).resolve().parents[1] / "spreadsheet_definitions" / "codex-spreadsheet-definition.md"


def _parse_definition_table(path: Path = DEFAULT_DEFINITION) -> Dict[str, Dict[str, bool]]:
    """Parse sheet-level rules from the markdown table."""
    rules: Dict[str, Dict[str, bool]] = {}
    if not path.exists():
        return rules
    for line in path.read_text().splitlines():
        if line.startswith("| CS_"):
            parts = [p.strip() for p in line.strip().strip("|").split("|")]
            if len(parts) >= 5:
                sheet = parts[0]
                req_map = parts[3].startswith("✅")
                has_formula = parts[4].startswith("✅")
                rules[sheet] = {
                    "requires_mapping": req_map,
                    "has_formula": has_formula,
                }
    return rules


def _str_series(df: pd.DataFrame, col: str) -> pd.Series:
    """Return a stripped string Series for ``col``, handling duplicate columns."""
    series = df[col]
    if isinstance(series, pd.DataFrame):
        series = series.iloc[:, 0]
    return series.astype(str).str.strip()


def validate_codeset_tab_logic(workbook_path: str | Path,
                               definition_path: Path | None = None) -> List[Dict[str, Any]]:
    """Validate mapping logic for each sheet in a workbook."""
    definition_path = definition_path or DEFAULT_DEFINITION
    rules = _parse_definition_table(definition_path)

    bytes_data = Path(workbook_path).read_bytes()
    try:
        xls = pd.ExcelFile(BytesIO(bytes_data), engine="openpyxl")
        wb = load_workbook(BytesIO(bytes_data), data_only=False)
    except ValueError:
        bytes_data = strip_invalid_font_families(bytes_data)
        xls = pd.ExcelFile(BytesIO(bytes_data), engine="openpyxl")
        wb = load_workbook(BytesIO(bytes_data), data_only=False)

    results: List[Dict[str, Any]] = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str).fillna("")
        ws = wb[sheet]

        cols = [c.upper().replace(" ", "_") for c in df.columns]
        std_code_col = None
        std_desc_col = None
        mapped_col = None
        code_col = None
        display_col = None
        for col, key in zip(df.columns, cols):
            if key == "STANDARD_CODE":
                std_code_col = col
            if key in [
                "STANDARD_DESCRIPTION",
                "STADARD_DESCRIPTION",
                "STANDARD_DESC",
                "STADARD_DESC",
            ]:
                std_desc_col = col
            if key in [
                "MAPPED_STD_DESCRIPTION",
                "MAPPED_STANDARD_DESCRIPTION",
                "MAPPED_STD_CODE",
                "MAPPED_STANDARD_CODE",
            ]:
                mapped_col = col
            if key == "CODE":
                code_col = col
            if key in ["DISPLAY_VALUE", "DISPLAY"]:
                display_col = col

        if code_col and display_col:
            mask = _str_series(df, code_col).ne("") | _str_series(df, display_col).ne("")
            df = df[mask]

        std_code_has = bool(std_code_col and _str_series(df, std_code_col).any())
        std_desc_has = bool(std_desc_col and _str_series(df, std_desc_col).any())
        requires_mapping = std_code_has and std_desc_has and mapped_col is not None

        missing = 0
        if requires_mapping and mapped_col:
            missing = int(_str_series(df, mapped_col).eq("").sum())

        sheet_rules = rules.get(sheet, {})
        formula_issues: List[str] = []
        if sheet_rules.get("has_formula"):
            cell = ws.cell(row=2, column=4)
            if cell.data_type != "f":
                formula_issues.append("Missing formula in D2")

        results.append({
            "sheet_name": sheet,
            "requires_mapping": requires_mapping,
            "missing_mapped_std_descriptions": missing,
            "formula_issues": formula_issues,
        })
    return results
