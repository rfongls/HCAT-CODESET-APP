"""Helper functions for exporting an updated workbook back to Excel.

The export routine operates on the original :class:`openpyxl.Workbook`
instance so cell styles, formulas, and validations remain intact. Only
cell values are replaced, preserving the workbook's formatting.
"""

from __future__ import annotations

from io import BufferedIOBase
from typing import Dict

import pandas as pd
from openpyxl.workbook.workbook import Workbook


def export_workbook(
    wb: Workbook,
    data: Dict[str, pd.DataFrame],
    stream: str | BufferedIOBase,
    protected: bool = False,
) -> None:
    """Write ``data`` back into ``wb`` and save it to ``stream``.

    Parameters
    ----------
    wb:
        The original workbook loaded via :mod:`openpyxl`. Formatting from
        this workbook is preserved while updating cell values.
    data:
        Mapping of sheet name to :class:`pandas.DataFrame` containing the
        latest cell values. Columns are matched to existing headers in the
        workbook and rows are written starting at row 2.
    stream:
        File path or binary file-like object to save the workbook to.
    protected:
        If ``True``, sheet protection is applied to every worksheet in the
        exported workbook.
    """

    for sheet, df in data.items():
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers)}

        # Prepare template styles from the first data row (row 2) if it exists
        style_row = 2 if ws.max_row >= 2 else None
        styles = {
            c_idx: ws.cell(row=style_row, column=c_idx)._style
            for c_idx in col_map.values()
        } if style_row else {}

        # Write DataFrame rows
        records = df.to_dict(orient="records")
        for r_idx, row in enumerate(records, start=2):
            for header, c_idx in col_map.items():
                value = row.get(header, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > ws.max_row and c_idx in styles:
                    cell._style = styles[c_idx]

        # Clear any remaining rows beyond the data frame length
        for r in range(len(records) + 2, ws.max_row + 1):
            for c_idx in col_map.values():
                ws.cell(row=r, column=c_idx, value=None)

    for ws in wb.worksheets:
        ws.protection.sheet = protected

    wb.save(stream)

