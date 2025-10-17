from __future__ import annotations
from typing import Dict, Any
from pathlib import Path
import tempfile
import io
from collections import defaultdict

import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, url_for
import json
from werkzeug.utils import secure_filename
from werkzeug.routing import BuildError
try:  # allow running as a package or standalone script
    from components.file_parser import load_workbook
    from components.dropdown_logic import extract_dropdown_options
    from components.formula_logic import extract_lookup_mappings
    from utils.export_excel import export_workbook
    from utils.transformer_xml import build_transformer_xml
    from validators import validate_workbook
except ModuleNotFoundError:  # pragma: no cover - fallback for imports when packaged
    from .components.file_parser import load_workbook
    from .components.dropdown_logic import extract_dropdown_options
    from .components.formula_logic import extract_lookup_mappings
    from .utils.export_excel import export_workbook
    from .utils.transformer_xml import build_transformer_xml
    from .validators import validate_workbook
from openpyxl.workbook.workbook import Workbook

app = Flask(__name__, static_folder="assets", template_folder="templates")
workbook_data: Dict[str, "pd.DataFrame"] = {}
dropdown_data: Dict[str, Dict[str, list]] = {}
mapping_data: Dict[str, Dict[str, Any]] = {}
field_notes: Dict[str, str] = {}
workbook_obj: Workbook | None = None
original_filename: str | None = None
workbook_path: Path | None = None
last_error: str | None = None
comparison_data: Dict[str, "pd.DataFrame"] = {}
comparison_path: Path | None = None
pending_import_path: Path | None = None
pending_import_name: str | None = None
pending_import_active: bool = False
pending_import_diff: Dict[str, Any] = {}

TRANSFORMER_REQUIRE_MAPPED = {
    "CS_ABNORMAL_FLAG",
    "CS_ADMIN_GENDER",
    "CS_ADMIT_SERVICE",
    "CS_ADMIT_SOURCE",
    "CS_ADMIT_TYPE",
    "CS_ALLERGY_REACTION_CODE",
    "CS_ALLERGY_SEVERITY_CODE",
    "CS_COMPLETION_STATUS",
    "CS_CONFIDENTIALITY_CODE",
    "CS_DIAGNOSIS_CODE_METHOD",
    "CS_DIAGNOSIS_TYPE",
    "CS_DIAGNOSTIC_SERVICE_SECTION",
    "CS_ENCOUNTER_CLASS",
    "CS_ETHNIC_GROUP",
    "CS_LANGUAGE",
    "CS_MARITAL_STATUS",
    "CS_ORDERING_PRIORITY",
    "CS_ORDER_STATUS",
    "CS_PROTECTION_IND",
    "CS_RACE",
    "CS_REL_TO_PERSON",
    "CS_RESULT_STATUS",
    "CS_RX_COMPONENT_TYPE",
    "CS_VIP_IND",
}

# File storing the user's preferred repository base path
CONFIG_FILE = Path(__file__).resolve().parent / "repo_base.txt"

# Directory containing sample repositories and workbooks. Defaults to the
# saved base path if available, otherwise the bundled Samples directory.
SAMPLES_DIR: Path | None = None


def _str_series(df: pd.DataFrame, col: str) -> pd.Series:
    """Return a stripped string Series for ``col`` selecting non-empty dupes."""
    series = df[col]
    if isinstance(series, pd.DataFrame):
        non_empty = series.ne("").sum()
        series = series.iloc[:, non_empty.values.argmax()]
    return series.astype(str).str.strip()


def _records(df: pd.DataFrame | None) -> list[dict]:
    """Convert ``df`` to record dicts, keeping the densest duplicate column."""
    if df is None:
        return []
    if df.columns.duplicated().any():
        cols = []
        for col in dict.fromkeys(df.columns):
            part = df.loc[:, df.columns == col]
            if isinstance(part, pd.Series):
                cols.append(part)
            else:
                non_empty = part.ne("").sum()
                cols.append(part.iloc[:, non_empty.values.argmax()])
        df = pd.concat(cols, axis=1)
        df.columns = list(dict.fromkeys(df.columns))
    return df.to_dict(orient="records")


def _clear_pending_import(clear_comparison: bool = True) -> None:
    """Remove any staged import workbook and reset comparison state if needed."""

    global pending_import_path, pending_import_name, pending_import_active
    global pending_import_diff, comparison_data, comparison_path

    staged_path = pending_import_path
    pending_import_path = None
    pending_import_name = None
    pending_import_diff = {}

    if pending_import_active and clear_comparison:
        comparison_data = {}
        comparison_path = None

    pending_import_active = False

    if staged_path and staged_path.exists():
        try:
            staged_path.unlink()
        except OSError:
            pass


def _infer_sheet_columns(df: pd.DataFrame | None) -> tuple[str | None, str | None, str | None]:
    """Infer key columns (code, display, mapped) from ``df`` when metadata is missing."""

    code_col = display_col = mapped_col = None
    if df is None:
        return code_col, display_col, mapped_col

    for col in df.columns:
        col_key = col.strip().upper().replace(" ", "_")
        if col_key == "CODE" and not code_col:
            code_col = col
        if col_key in {"DISPLAY_VALUE", "DISPLAY"} and not display_col:
            display_col = col
        if col_key in {
            "MAPPED_STD_DESCRIPTION",
            "MAPPED_STANDARD_DESCRIPTION",
            "MAPPED_STD_CODE",
            "MAPPED_STANDARD_CODE",
        } and not mapped_col:
            mapped_col = col
    return code_col, display_col, mapped_col


def _normalize_diff_dataframe(df: pd.DataFrame | None, columns: list[str]) -> pd.DataFrame:
    """Return a normalized dataframe limited to ``columns`` for diff comparisons."""

    if not columns:
        return pd.DataFrame()

    if df is None or df.empty:
        return pd.DataFrame({col: pd.Series(dtype=str) for col in columns})

    data: Dict[str, pd.Series] = {}
    for col in columns:
        if col in df.columns:
            series = _str_series(df, col)
        else:
            series = pd.Series([""] * len(df))
        data[col] = series.fillna("").astype(str).str.strip()
    return pd.DataFrame(data).reset_index(drop=True)


def _prepare_diff_rows(
    sheet: str,
    df: pd.DataFrame | None,
    columns: list[str],
    key_column: str | None,
) -> list[Dict[str, Any]]:
    """Convert dataframe rows into a diff-friendly structure with unique keys."""

    normalized = _normalize_diff_dataframe(df, columns)
    rows: list[Dict[str, Any]] = []

    if normalized.empty:
        return rows

    for idx, record in enumerate(normalized.to_dict(orient="records")):
        row_number = idx + 2  # account for header row in Excel
        key_value = record.get(key_column, "") if key_column else ""
        display_key = key_value or f"Row {row_number}"
        rows.append(
            {
                "values": record,
                "row": row_number,
                "key_value": key_value.strip(),
                "display_key": display_key,
            }
        )

    occurrences: dict[str, int] = defaultdict(int)
    for row in rows:
        base_key = row["key_value"] or f"__row_{row['row']}"
        count = occurrences[base_key]
        occurrences[base_key] += 1
        unique_key = base_key if count == 0 else f"{base_key}#{count}"
        row["unique_key"] = unique_key

    return rows


def _compute_workbook_diff(imported: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
    """Build a diff summary between the active workbook and ``imported`` data."""

    diff: Dict[str, Any] = {"sheets": {}, "summary": {}}
    sheet_names = sorted(set(workbook_data.keys()) | set(imported.keys()))

    total_added = total_removed = total_changed = 0

    for sheet in sheet_names:
        base_df = workbook_data.get(sheet)
        new_df = imported.get(sheet)

        info = mapping_data.get(sheet, {})
        base_code, base_display, base_mapped = _infer_sheet_columns(base_df)
        new_code, new_display, new_mapped = _infer_sheet_columns(new_df)

        def _pick_column(candidates: list[str | None]) -> str | None:
            for cand in candidates:
                if not cand:
                    continue
                if (base_df is not None and cand in base_df.columns) or (
                    new_df is not None and cand in new_df.columns
                ):
                    return cand
            return None

        code_col = _pick_column([info.get("code_col"), base_code, new_code])
        display_col = _pick_column([info.get("display_col"), base_display, new_display])
        mapped_col = _pick_column([info.get("mapped_col"), base_mapped, new_mapped])

        columns = [c for c in [code_col, display_col, mapped_col] if c]
        if not columns:
            continue

        columns = list(dict.fromkeys(columns))
        key_column = code_col or display_col or mapped_col

        base_rows = _prepare_diff_rows(sheet, base_df, columns, key_column)
        new_rows = _prepare_diff_rows(sheet, new_df, columns, key_column)

        base_map = {row["unique_key"]: row for row in base_rows}
        new_map = {row["unique_key"]: row for row in new_rows}

        added_rows: list[Dict[str, Any]] = []
        removed_rows: list[Dict[str, Any]] = []
        changed_rows: list[Dict[str, Any]] = []

        for key, new_row in new_map.items():
            if key not in base_map:
                added_rows.append(
                    {
                        "key": new_row["display_key"],
                        "values": new_row["values"],
                        "row": new_row["row"],
                    }
                )
            else:
                base_row = base_map[key]
                if base_row["values"] != new_row["values"]:
                    changed_columns = [
                        col for col in columns if base_row["values"].get(col) != new_row["values"].get(col)
                    ]
                    changed_rows.append(
                        {
                            "key": new_row["display_key"],
                            "before": base_row["values"],
                            "after": new_row["values"],
                            "base_row": base_row["row"],
                            "incoming_row": new_row["row"],
                            "changed_columns": changed_columns,
                        }
                    )

        for key, base_row in base_map.items():
            if key not in new_map:
                removed_rows.append(
                    {
                        "key": base_row["display_key"],
                        "values": base_row["values"],
                        "row": base_row["row"],
                    }
                )

        if added_rows or removed_rows or changed_rows:
            diff["sheets"][sheet] = {
                "columns": columns,
                "key_column": key_column,
                "added": added_rows,
                "removed": removed_rows,
                "changed": changed_rows,
                "added_count": len(added_rows),
                "removed_count": len(removed_rows),
                "changed_count": len(changed_rows),
            }
            total_added += len(added_rows)
            total_removed += len(removed_rows)
            total_changed += len(changed_rows)

    diff["summary"] = {
        "total_sheets": len(diff["sheets"]),
        "added": total_added,
        "removed": total_removed,
        "changed": total_changed,
    }
    diff["has_changes"] = bool(diff["sheets"])
    return diff


def _stage_import_workbook(path: Path, filename: str) -> Dict[str, Any]:
    """Load ``path`` as a staged import workbook and populate comparison data."""

    global pending_import_path, pending_import_name, pending_import_active, pending_import_diff

    with path.open("rb") as fh:
        imported_data, _ = load_workbook(fh)

    pending_import_diff = _compute_workbook_diff(imported_data)
    pending_import_path = path
    pending_import_name = filename
    pending_import_active = True

    _load_comparison_workbook_path(path)
    return pending_import_diff


def load_repository_base() -> None:
    """Load repository base from config and refresh cache."""
    global SAMPLES_DIR
    if CONFIG_FILE.exists():
        SAMPLES_DIR = Path(CONFIG_FILE.read_text().strip())
    else:
        SAMPLES_DIR = Path(__file__).resolve().parent.parent / "Samples"
    refresh_repository_cache()

def discover_repository_workbooks(base: Path) -> Dict[str, list[str]]:
    """Return a mapping of repository *relative paths* to codeset workbooks.

    The scan searches for any directory containing ``Codeset`` in its name and
    looks for ``*.xlsx`` files with ``Codeset`` in the filename. For each file
    found, the nearest ancestor directory whose name contains ``Repository`` or
    ends with ``-prc`` (case-insensitive) is treated as the repository root. If
    no such ancestor is found, the file is grouped under a pseudo repository named
    ``SharedRepositories``. The returned mapping uses paths relative to the
    repository root (or ``base`` for shared files). Repository keys are stored as
    their path relative to ``base`` to avoid collisions between repositories
    with the same name and to allow nested repositories to be addressed
    correctly.
    """

    repo_map: Dict[str, list[str]] = {}
    shared_files: list[str] = []
    if base and base.exists():
        for file in base.rglob("*.xlsx"):
            if "codeset" not in file.name.lower():
                continue
            repo_dir: Path | None = None
            for ancestor in file.parents:
                if ancestor == base:
                    break
                name = ancestor.name.lower()
                if "repository" in name or name.endswith("-prc"):
                    # keep the first (closest) repository ancestor
                    repo_dir = ancestor
                    break
            if repo_dir is not None:
                rel_repo = str(repo_dir.relative_to(base))
                rel_path = str(file.relative_to(repo_dir))
                repo_map.setdefault(rel_repo, []).append(rel_path)
            else:
                shared_files.append(str(file.relative_to(base)))

    for name, files in repo_map.items():
        repo_map[name] = sorted(set(files))
    if shared_files:
        repo_map["SharedRepositories"] = sorted(set(shared_files))
    return repo_map


# Cached mapping of repositories to workbooks for the currently selected base
REPOSITORY_CACHE: Dict[str, list[str]] = {}


def refresh_repository_cache() -> None:
    """Refresh repository cache; used at startup or when Samples path changes."""
    global REPOSITORY_CACHE
    REPOSITORY_CACHE = (
        discover_repository_workbooks(SAMPLES_DIR)
        if SAMPLES_DIR is not None
        else {}
    )

load_repository_base()


def _load_workbook_path(path: Path, filename: str) -> None:
    """Load workbook at ``path`` and populate globals for UI rendering."""
    global workbook_data, workbook_obj, dropdown_data, mapping_data, field_notes, original_filename, last_error, comparison_data, comparison_path

    _clear_pending_import(clear_comparison=False)

    with path.open("rb") as fh:
        workbook_data, wb = load_workbook(fh)
    workbook_obj = wb
    original_filename = filename
    dropdown_data = extract_dropdown_options(wb)
    lookup_maps = extract_lookup_mappings(wb)

    mapping_data = {}
    field_notes = {}

    for sheet, df in list(workbook_data.items()):
        mapped_col = None
        mapped_type: str | None = None  # "description" or "code"
        sub_col = None
        std_col = None
        std_code_col = None
        code_col = None
        display_col = None
        hidden_cols: list[str] = []
        definition_col = None
        for col in df.columns:
            col_key = col.strip().upper().replace(" ", "_")
            if col_key in ["MAPPED_STANDARD_DESCRIPTION", "MAPPED_STD_DESCRIPTION"]:
                mapped_col = col
                mapped_type = "description"
            if col_key in ["MAPPED_STANDARD_CODE", "MAPPED_STD_CODE"]:
                mapped_col = col
                mapped_type = "code"
            if col_key in [
                "SUB_DEFINITION",
                "SUB_DEFINITION_DESCRIPTION",
                "SUBDEFINITION",
                "SUB DEFINITION",
            ]:
                sub_col = col
            if col_key in [
                "STANDARD_DESCRIPTION",
                "STD_DESCRIPTION",
                "STANDARD_DESC",
                "STADARD_DESCRIPTION",
                "STADARD_DESC",
            ]:
                std_col = col
            if col_key in ["STANDARD_CODE", "STD_CODE"]:
                std_code_col = col
            if col_key == "CODE":
                code_col = col
            if col_key in ["DISPLAY_VALUE", "DISPLAY"]:
                display_col = col
            if col_key == "DEFINITION":
                definition_col = col
                hidden_cols.append(col)

        if mapped_col is None and std_col is None and std_code_col is None and definition_col:
            mapped_col = definition_col
            mapped_type = "description"
            hidden_cols = [c for c in hidden_cols if c != mapped_col]

        if mapped_col:
            source_col = std_col if mapped_type != "code" else std_code_col
            if source_col is None:
                source_col = mapped_col
            options = sorted({v for v in _str_series(df, source_col) if v})
            sheet_opts = dropdown_data.setdefault(sheet, {})
            if options and mapped_col not in sheet_opts:
                sheet_opts[mapped_col] = options

        sheet_map: Dict[str, str] = {}

        if std_col and std_code_col:
            std_series = _str_series(df, std_col)
            code_series = _str_series(df, std_code_col)
            mask = std_series != ""
            if mapped_type == "code":
                sheet_map.update({code: f"{code}^{desc}" for code, desc in zip(code_series[mask], std_series[mask])})
            else:
                sheet_map.update({desc: f"{code}^{desc}" for desc, code in zip(std_series[mask], code_series[mask])})

        if mapped_col and not (std_col and std_code_col) and sub_col:
            mapped_series = _str_series(df, mapped_col)
            sub_series = _str_series(df, sub_col)
            mask = mapped_series != ""
            sheet_map.update({k: v for k, v in zip(mapped_series[mask], sub_series[mask])})

        lookup_sheet = lookup_maps.get(sheet, {})
        if sub_col and sub_col in lookup_sheet:
            sheet_map = {**lookup_sheet[sub_col], **sheet_map}

        if sub_col and mapped_col:
            df[sub_col] = df[mapped_col].map(sheet_map).fillna(df[sub_col])

        mapping_data[sheet] = {
            "map": sheet_map,
            "sub_col": sub_col,
            "mapped_col": mapped_col,
            "code_col": code_col,
            "display_col": display_col,
            "std_col": std_col,
            "std_code_col": std_code_col,
            "hidden_cols": hidden_cols,
        }

        if code_col and display_col and mapped_col:
            code_series = _str_series(df, code_col)
            display_series = _str_series(df, display_col)
            blank_mask = code_series.eq("") & display_series.eq("")
            if blank_mask.any():
                df.loc[blank_mask, mapped_col] = ""
                if sub_col:
                    df.loc[blank_mask, sub_col] = ""
        workbook_data[sheet] = df

        note = ""
        if code_col:
            ws = wb[sheet]
            for cell in ws[1]:
                if (cell.value or "").strip() == code_col and cell.comment:
                    note = cell.comment.text.strip()
                    break
        field_notes[sheet] = note

    comparison_data = {}
    comparison_path = None
    last_error = None


def _load_comparison_workbook_path(path: Path) -> None:
    """Load a comparison workbook keeping only key columns."""
    global comparison_data, comparison_path
    comparison_data = {}
    with path.open("rb") as fh:
        data, _ = load_workbook(fh)
    for sheet, df in data.items():
        code_col = None
        display_col = None
        mapped_col = None
        for col in df.columns:
            col_key = col.strip().upper().replace(" ", "_")
            if col_key == "CODE":
                code_col = col
            if col_key in ["DISPLAY_VALUE", "DISPLAY"]:
                display_col = col
            if col_key in [
                "MAPPED_STD_DESCRIPTION",
                "MAPPED_STANDARD_DESCRIPTION",
                "MAPPED_STD_CODE",
                "MAPPED_STANDARD_CODE",
            ]:
                mapped_col = col
        cols: Dict[str, Any] = {}
        if code_col:
            cols["CODE_COMPARE"] = df[code_col]
        if display_col:
            cols["DISPLAY_VALUE_COMPARE"] = df[display_col]
        if mapped_col:
            cols[f"{mapped_col}_COMPARE"] = df[mapped_col]
        if cols:
            comparison_data[sheet] = pd.DataFrame(cols)
    comparison_path = path



def _combine_sheet(sheet: str) -> pd.DataFrame | None:
    """Return sheet data with comparison columns merged in."""
    df = workbook_data.get(sheet)
    if df is None:
        return None
    cmp = comparison_data.get(sheet)
    if cmp is None:
        return df
    combined = df.copy()
    info = mapping_data.get(sheet, {})
    code_col = info.get("code_col")
    display_col = info.get("display_col")
    mapped_col = info.get("mapped_col")
    if code_col and "CODE_COMPARE" in cmp:
        combined.insert(combined.columns.get_loc(code_col) + 1, "CODE_COMPARE", cmp["CODE_COMPARE"])
    if display_col and "DISPLAY_VALUE_COMPARE" in cmp:
        combined.insert(
            combined.columns.get_loc(display_col) + 1,
            "DISPLAY_VALUE_COMPARE",
            cmp["DISPLAY_VALUE_COMPARE"],
        )
    if mapped_col:
        cmp_key = f"{mapped_col}_COMPARE"
        if cmp_key in cmp:
            combined.insert(
                combined.columns.get_loc(mapped_col) + 1,
                cmp_key,
                cmp[cmp_key],
            )
    return combined

@app.route("/", methods=["GET", "POST"])
def index():
    global workbook_data
    global dropdown_data
    global last_error
    global mapping_data
    global workbook_obj
    global original_filename
    global workbook_path, comparison_path
    global SAMPLES_DIR
    selected_repo: str | None = None
    selected_workbook: str | None = None
    selected_compare_repo: str | None = None
    selected_compare_workbook: str | None = None
    repo_names = sorted(REPOSITORY_CACHE.keys())
    repo_display = {n: (n if n == "SharedRepositories" else Path(n).name) for n in repo_names}
    repo_files: list[str] = []
    compare_repo_files: list[str] = []
    compare_mode = bool(comparison_data)
    reopen_controls = False

    if request.method == "POST":
        reopen_controls = bool(request.form.get("open_overlay"))
        if "repo_base" in request.form:
            # user provided a parent directory to scan for repositories
            base_path = Path(request.form.get("repo_base", "")).expanduser()
            if base_path.is_dir():
                SAMPLES_DIR = base_path
                if request.form.get("save_repo_base"):
                    CONFIG_FILE.write_text(str(base_path))
                refresh_repository_cache()
                repo_names = sorted(REPOSITORY_CACHE.keys())
                repo_display = {
                    n: (n if n == "SharedRepositories" else Path(n).name)
                    for n in repo_names
                }
            else:
                last_error = "Repository folder not found"
        elif request.form.get("end_compare"):
            comparison_data.clear()
            comparison_path = None
            compare_mode = False
        elif request.form.get("start_compare"):
            compare_mode = True
        elif "workbook" in request.files and not request.form.get("compare_mode"):
            file = request.files["workbook"]
            if file.filename:
                try:
                    filename = secure_filename(file.filename)
                    temp_dir = Path(tempfile.gettempdir())
                    workbook_path = temp_dir / filename
                    file.save(workbook_path)
                    _load_workbook_path(workbook_path, filename)
                except Exception as exc:
                    last_error = str(exc)
                    workbook_data = {}
                    dropdown_data = {}
                    mapping_data = {}
        elif request.form.get("repo") and request.form.get("workbook_name"):
            selected_repo = request.form.get("repo")
            selected_workbook = request.form.get("workbook_name")
            try:
                if SAMPLES_DIR is None:
                    raise FileNotFoundError("Repository folder not selected")
                base = SAMPLES_DIR.resolve()
                if selected_repo == "SharedRepositories":
                    workbook_path = (SAMPLES_DIR / selected_workbook).resolve()
                    if not workbook_path.is_file() or not workbook_path.is_relative_to(base):
                        raise FileNotFoundError("Workbook not found")
                else:
                    repo_path = (SAMPLES_DIR / selected_repo).resolve()
                    if not repo_path.is_dir() or not repo_path.is_relative_to(base):
                        raise FileNotFoundError("Repository not found")
                    workbook_path = repo_path / selected_workbook
                    if not workbook_path.is_file():
                        raise FileNotFoundError("Workbook not found")
                _load_workbook_path(workbook_path, selected_workbook)
            except Exception as exc:
                last_error = str(exc)
                workbook_data = {}
                dropdown_data = {}
                mapping_data = {}
        elif request.form.get("compare_repo") and request.form.get("compare_workbook_name"):
            try:
                repo = request.form.get("compare_repo")
                wb_name = request.form.get("compare_workbook_name")
                if SAMPLES_DIR is None:
                    raise FileNotFoundError("Repository folder not selected")
                base = SAMPLES_DIR.resolve()
                repo_path = (SAMPLES_DIR / repo).resolve()
                if not repo_path.is_dir() or not repo_path.is_relative_to(base):
                    raise FileNotFoundError("Repository not found")
                cmp_path = repo_path / wb_name
                if not cmp_path.is_file():
                    raise FileNotFoundError("Workbook not found")
                _load_comparison_workbook_path(cmp_path)
                compare_mode = True
            except Exception as exc:
                last_error = str(exc)
                comparison_data.clear()

    if workbook_path and not selected_repo:
        try:
            if SAMPLES_DIR and workbook_path.is_relative_to(SAMPLES_DIR):
                selected_repo = str(workbook_path.parent.relative_to(SAMPLES_DIR))
                selected_workbook = workbook_path.name
            else:
                selected_repo = workbook_path.parent.name
                selected_workbook = workbook_path.name
        except Exception:
            selected_repo = selected_workbook = None

    if comparison_path:
        try:
            if SAMPLES_DIR and comparison_path.is_relative_to(SAMPLES_DIR):
                selected_compare_repo = str(
                    comparison_path.parent.relative_to(SAMPLES_DIR)
                )
                selected_compare_workbook = comparison_path.name
                compare_repo_files = REPOSITORY_CACHE.get(selected_compare_repo, [])
            else:
                selected_compare_repo = comparison_path.parent.name
                selected_compare_workbook = comparison_path.name
                compare_repo_files = []
        except Exception:
            selected_compare_repo = selected_compare_workbook = None
            compare_repo_files = []

    if selected_repo:
        repo_files = REPOSITORY_CACHE.get(selected_repo, [])

    sheet_names = list(workbook_data.keys())
    base_headers: Dict[str, list] = {s: df.columns.tolist() for s, df in workbook_data.items()}
    render_headers: Dict[str, list] = {}
    for s in sheet_names:
        cols = _combine_sheet(s).columns.tolist()
        hidden = mapping_data.get(s, {}).get("hidden_cols", [])
        cols = [c for c in cols if c not in hidden]
        render_headers[s] = cols
    initial_sheet = sheet_names[0] if sheet_names else None
    initial_records = _records(_combine_sheet(initial_sheet)) if initial_sheet else []
    base_initial_records = _records(workbook_data.get(initial_sheet)) if initial_sheet else []
    initial_compare_records = (
        _records(comparison_data.get(initial_sheet, pd.DataFrame())) if initial_sheet else []
    )
    comparison_repos = [r for r in repo_names if r != selected_repo]
    try:
        transformer_url = url_for("export_transformer")
    except BuildError:
        transformer_url = None

    initial_errors: list[str] = []
    if workbook_data:
        try:
            initial_errors = validate_workbook(workbook_data, mapping_data)
        except Exception:
            initial_errors = []

    return render_template(
        "index.html",
        sheet_names=sheet_names,
        initial_sheet=initial_sheet,
        initial_records=initial_records,
        base_records=base_initial_records,
        compare_records=initial_compare_records,
        headers=base_headers,
        render_headers=render_headers,
        dropdowns=dropdown_data,
        mappings=mapping_data,
        field_notes=field_notes,
        error=last_error,
        filename=original_filename,
        repositories=repo_names,
        repo_display=repo_display,
        repo_base=SAMPLES_DIR,
        comparison_repositories=comparison_repos,
        repo_files=repo_files,
        selected_repo=selected_repo,
        selected_workbook=selected_workbook,
        selected_compare_repo=selected_compare_repo,
        selected_compare_workbook=selected_compare_workbook,
        compare_repo_files=compare_repo_files,
        comparison_active=bool(comparison_data),
        compare_mode=compare_mode,
        reopen_controls=reopen_controls,
        transformer_url=transformer_url,
        initial_errors=initial_errors,
        pending_import=pending_import_diff,
        pending_import_active=pending_import_active,
        pending_import_name=pending_import_name,
    )


@app.route("/sheet/<sheet_name>", endpoint="sheet_data")
def sheet_data(sheet_name: str):
    df = _combine_sheet(sheet_name)
    if df is None:
        return jsonify([])
    return jsonify(_records(df))


@app.route("/workbooks/<path:repo>")
def list_workbooks(repo: str):
    """Return available workbooks for ``repo`` from the cached scan."""
    return jsonify(REPOSITORY_CACHE.get(repo, []))


@app.route("/transformer")
def export_transformer():
    """Generate an XML transformer from the currently loaded workbook."""
    global workbook_data
    if not workbook_data:
        return "No workbook loaded", 400

    skip_mapped_requirement = {
        sheet
        for sheet in workbook_data
        if sheet not in TRANSFORMER_REQUIRE_MAPPED
    }
    errors = validate_workbook(
        workbook_data,
        mapping_data,
        skip_mapped_requirement_sheets=skip_mapped_requirement,
    )
    if errors:
        return jsonify({"errors": errors}), 400

    free_param = request.args.get("freetext")
    free_map = {}
    if free_param:
        try:
            free_map = json.loads(free_param)
        except json.JSONDecodeError:
            free_map = {}

    xml_str = build_transformer_xml(workbook_data, free_map)
    return send_file(
        io.BytesIO(xml_str.encode("utf-8")),
        mimetype="application/xml",
        as_attachment=True,
        download_name="CodesetTransformer.xml",
    )


@app.route("/export", methods=["POST"])
def export():
    """Export the in-memory workbook with updated values."""
    global workbook_obj, workbook_data, original_filename, workbook_path
    if workbook_obj is None or workbook_path is None:
        return "No workbook loaded", 400

    payload = request.get_json() or {}
    if not isinstance(payload, dict):
        return "Invalid payload", 400

    workbook_payload = payload.get("data") if "data" in payload else payload
    locks = payload.get("locks", False)
    if not isinstance(workbook_payload, dict) or not isinstance(locks, bool):
        return "Invalid payload", 400

    for sheet, rows in workbook_payload.items():
        if sheet in workbook_data:
            df = pd.DataFrame(rows, columns=workbook_data[sheet].columns)
            df = df.where(pd.notna(df), "")
            workbook_data[sheet] = df

    errors = validate_workbook(workbook_data, mapping_data)
    if errors:
        return jsonify({"errors": errors}), 400

    # Write to a temporary file and atomically replace the original so the
    # on-disk workbook is always updated in place
    tmp_path = workbook_path.with_name(workbook_path.name + ".tmp")
    export_workbook(workbook_obj, workbook_data, tmp_path, locks)
    tmp_path.replace(workbook_path)

    filename = original_filename or workbook_path.name
    return jsonify({"status": "ok", "filename": filename})


@app.route("/export_errors", methods=["POST"])
def export_errors():
    """Return a CSV file listing validation errors for the current data."""
    global workbook_data, workbook_obj, workbook_path
    if workbook_obj is None or workbook_path is None:
        return "No workbook loaded", 400

    payload = request.get_json() or {}
    if not isinstance(payload, dict):
        return "Invalid payload", 400

    workbook_payload = payload.get("data") if "data" in payload else payload
    if not isinstance(workbook_payload, dict):
        return "Invalid payload", 400

    for sheet, rows in workbook_payload.items():
        if sheet in workbook_data:
            df = pd.DataFrame(rows, columns=workbook_data[sheet].columns)
            df = df.where(pd.notna(df), "")
            workbook_data[sheet] = df

    errors = validate_workbook(workbook_data, mapping_data)
    if not errors:
        return jsonify({"errors": []})

    buf = io.StringIO()
    pd.DataFrame({"Error": errors}).to_csv(buf, index=False)
    mem = io.BytesIO(buf.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name="error_report.csv",
        mimetype="text/csv",
    )

@app.route("/import", methods=["POST"])
def import_workbook():
    """Overwrite the current workbook with an uploaded file and reload."""
    global workbook_path
    if workbook_path is None:
        return "No workbook loaded", 400
    file = request.files.get("workbook")
    if file is None or not file.filename:
        return "No workbook provided", 400
    stage_only = request.form.get("stage_only")
    tmp_path: Path | None = None
    try:
        filename = secure_filename(file.filename)
        suffix = Path(filename).suffix or ".xlsx"
        _clear_pending_import()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp_path = Path(tmp.name)
        tmp.close()
        file.save(tmp_path)
        if stage_only:
            diff = _stage_import_workbook(tmp_path, filename)
            status = "pending_changes" if diff.get("has_changes") else "pending_no_changes"
            return jsonify({"status": status, "diff": diff})
        tmp_path.replace(workbook_path)
        tmp_path = None
        _load_workbook_path(workbook_path, filename or workbook_path.name)
        return jsonify({"status": "applied"})
    except Exception as exc:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        _clear_pending_import()
        return str(exc), 400


@app.route("/import/confirm", methods=["POST"])
def confirm_import_workbook():
    """Apply or discard the currently staged import workbook."""

    global workbook_path, pending_import_path, pending_import_name

    if workbook_path is None:
        return "No workbook loaded", 400

    payload = request.get_json(silent=True) or {}
    action = payload.get("action")

    if action == "cancel":
        _clear_pending_import()
        return jsonify({"status": "cancelled"})

    if action != "apply_all":
        return "Invalid action", 400

    staged_path = pending_import_path
    if staged_path is None or not staged_path.exists():
        return "No pending import", 400

    try:
        staged_path.replace(workbook_path)
        _load_workbook_path(workbook_path, pending_import_name or workbook_path.name)
    except Exception as exc:
        _clear_pending_import()
        return str(exc), 400

    _clear_pending_import()
    return jsonify({"status": "applied"})

def _first_non_loopback_ipv4() -> str:
    """Return the first non-loopback IPv4 address for the current host."""
    import socket

    try:
        hostname = socket.gethostname()
        for addr in socket.gethostbyname_ex(hostname)[2]:
            if not addr.startswith("127.") and not addr.startswith("169.254"):
                return addr
    except socket.gaierror:
        pass

    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            return sock.getsockname()[0]
    except OSError:
        pass

    return "127.0.0.1"


def _launch_browser_when_ready(port: int) -> None:
    """Open the default browser once the development server is reachable."""
    import socket
    import threading
    import time
    import webbrowser

    def _open() -> None:
        deadline = time.time() + 40
        while time.time() < deadline:
            try:
                with socket.create_connection(("127.0.0.1", port), timeout=1):
                    break
            except OSError:
                time.sleep(0.5)
        host_ip = _first_non_loopback_ipv4()
        url = f"http://{host_ip}:{port}/"
        try:
            webbrowser.open(url)
        except Exception:
            # Fallback to printing the URL so the operator can open it manually.
            print(f"Codeset UI available at {url}")

    threading.Thread(target=_open, daemon=True).start()


if __name__ == "__main__":
    # Disable the Flask reloader so the server doesn't restart when a workbook
    # is loaded. The reloader can interrupt the initial request and appear as a
    # connection reset in the browser.
    port = 5000
    _launch_browser_when_ready(port)
    app.run(host="0.0.0.0", port=port, debug=True, use_reloader=False)
