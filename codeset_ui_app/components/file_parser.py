from __future__ import annotations
from typing import Dict, Tuple
from io import BytesIO
from ..utils.xlsx_sanitizer import strip_invalid_font_families

# ``file_parser`` can be imported either as part of the ``codeset_ui_app``
# package (via unit tests) or executed when running ``app.py`` directly. Use a
# flexible import so the sanitizer is found in both contexts.
try:  # pragma: no cover - import resolution path tested indirectly
    from ..utils.xlsx_sanitizer import strip_invalid_font_families
except ImportError:  # running as a script from the ``codeset_ui_app`` directory
    from utils.xlsx_sanitizer import strip_invalid_font_families

import pandas as pd
from openpyxl import load_workbook as _load_workbook
from openpyxl.workbook.workbook import Workbook


def load_workbook(file) -> Tuple[Dict[str, pd.DataFrame], Workbook]:
    """Return workbook data and an openpyxl workbook instance.

    The workbook bytes are read once and reused for both :mod:`pandas`
    (to obtain values with formulas resolved) and :mod:`openpyxl` (to
    inspect formulas and validations).
    """
    file_bytes = file.read()
    try:
        data_wb = _load_workbook(BytesIO(file_bytes), data_only=True)
        wb = _load_workbook(BytesIO(file_bytes), data_only=False)
    except ValueError:
        file_bytes = strip_invalid_font_families(file_bytes)
        data_wb = _load_workbook(BytesIO(file_bytes), data_only=True)
        wb = _load_workbook(BytesIO(file_bytes), data_only=False)

    data: Dict[str, pd.DataFrame] = {}
    for sheet in data_wb.sheetnames:
        ws = data_wb[sheet]
        rows = list(ws.values)
        if not rows:
            data[sheet] = pd.DataFrame()
            continue
        header, *content = rows
        df = pd.DataFrame(content, columns=header)
        df = df.where(pd.notna(df), "").astype(str)
        # Remove completely empty rows but preserve blank columns so the UI retains
        # all expected headers, even when no data is present in a column.
        df.replace("", pd.NA, inplace=True)
        df.dropna(axis=0, how="all", inplace=True)
        df.fillna("", inplace=True)
        empty_cols = [c for c in df.columns if not c or str(c).startswith("Unnamed")]
        df.drop(columns=empty_cols, inplace=True, errors="ignore")
        data[sheet] = df

    return data, wb
