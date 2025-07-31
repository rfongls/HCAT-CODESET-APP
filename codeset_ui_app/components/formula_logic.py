from __future__ import annotations

from typing import Dict, Any
import re
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def extract_column_formulas(file) -> Dict[str, Dict[str, str]]:
    """Return formulas for the first data row of each column per sheet."""
    file.seek(0)
    wb = load_workbook(file, data_only=False)
    formulas: Dict[str, Dict[str, str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
        sheet_formulas: Dict[str, str] = {}
        # Inspect row 2 which typically holds the first data row
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=False):
            for header, cell in zip(headers, row):
                if cell.data_type == "f" and isinstance(cell.value, str):
                    sheet_formulas[header] = cell.value
            break
        formulas[sheet_name] = sheet_formulas
    return formulas


def _parse_vlookup_range(formula: str, wb) -> Dict[str, str] | None:
    """Return mapping from the table referenced in a VLOOKUP formula.

    Only handles simple formulas like ``VLOOKUP(A2,Sheet!A:B,2,0)``. Named ranges
    are also supported. If parsing fails, ``None`` is returned.
    """
    m = re.search(r"VLOOKUP\([^,]+,([^,]+),(\d+)", formula, re.IGNORECASE)
    if not m:
        return None
    table_ref = m.group(1).strip().strip('"')
    try:
        col_index = int(m.group(2))
    except Exception:
        return None

    if "!" in table_ref:
        sheet_name, cell_range = table_ref.split("!", 1)
        sheet_name = sheet_name.strip("'")
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
    else:
        if table_ref in wb.defined_names:
            defn = wb.defined_names[table_ref]
            dest = next(iter(defn.destinations), (None, None))
            if dest[0] is None:
                return None
            ws = wb[dest[0]]
            cell_range = dest[1]
        else:
            ws = wb.active
            cell_range = table_ref

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except Exception:
        return None

    mapping: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        key_cell = row[0]
        value_cell = row[col_index - 1] if col_index - 1 < len(row) else None
        key = key_cell.value
        val = value_cell.value if value_cell else None
        if key is not None:
            mapping[str(key)] = "" if val is None else str(val)
    return mapping


def extract_lookup_mappings(file) -> Dict[str, Dict[str, Dict[str, str]]]:
    """Extract lookup mappings defined via simple VLOOKUP formulas."""
    file.seek(0)
    wb = load_workbook(file, data_only=False)
    mappings: Dict[str, Dict[str, Dict[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
        sheet_map: Dict[str, Dict[str, str]] = {}
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=False):
            for header, cell in zip(headers, row):
                if cell.data_type == "f" and isinstance(cell.value, str):
                    mapping = _parse_vlookup_range(cell.value, wb)
                    if mapping:
                        sheet_map[header] = mapping
            break
        if sheet_map:
            mappings[sheet_name] = sheet_map
    return mappings
