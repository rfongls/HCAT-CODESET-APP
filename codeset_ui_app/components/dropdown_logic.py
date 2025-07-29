"""Utilities for extracting and handling dropdown validations."""

from __future__ import annotations

from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def _parse_formula(formula: str, wb) -> List[str]:
    """Return a list of options from a data-validation formula.

    Any errors in the formula (for example, ``#REF!`` references) result
    in an empty list rather than an exception.
    """
    if not formula:
        return []
    formula = formula.lstrip("=")
    try:
        # Quoted comma-separated list
        if formula.startswith('"') and formula.endswith('"'):
            return [v.strip() for v in formula.strip('"').split(',') if v.strip()]

        # Named range reference
        if formula in wb.defined_names:
            values: List[str] = []
            defn = wb.defined_names[formula]
            for title, coord in defn.destinations:
                try:
                    sheet = wb[title]
                    min_col, min_row, max_col, max_row = range_boundaries(coord)
                except Exception:
                    continue
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        value = sheet.cell(row=row, column=col).value
                        if value is not None:
                            values.append(str(value))
            return values
        # Range reference, e.g. Sheet1!$A$1:$A$5
        if "!" in formula:
            sheet_name, cell_range = formula.split("!", 1)
            if sheet_name == "#REF":
                return []
            try:
                sheet = wb[sheet_name]
            except KeyError:
                return []
        else:
            # Current sheet range
            cell_range = formula
            sheet = wb.active

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        values: List[str] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    values.append(str(value))
        return values
    except Exception:
        # Gracefully handle malformed formulas
        return []


def extract_dropdown_options(file) -> Dict[str, Dict[str, List[str]]]:
    """Extract dropdown validation options per sheet and column.

    Parameters
    ----------
    file: path or file-like object
        The Excel workbook to inspect.

    Returns
    -------
    Dict[str, Dict[str, List[str]]]
        Mapping of sheet names to columns and their list of allowed values.
    """
    file.seek(0)
    wb = load_workbook(file, data_only=True)
    dropdowns: Dict[str, Dict[str, List[str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_opts: Dict[str, List[str]] = {}
        headers = {cell.column: str(cell.value) for cell in ws[1]}
        if ws.data_validations is None:
            dropdowns[sheet_name] = sheet_opts
            continue
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue
            options = _parse_formula(dv.formula1, wb)
            for rng in dv.cells.ranges:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(str(rng))
                except Exception:
                    continue
                # assume validation applies to entire column below header
                if min_row <= 2:
                    for col in range(min_col, max_col + 1):
                        header = headers.get(col)
                        if header:
                            sheet_opts.setdefault(header, options)
        dropdowns[sheet_name] = sheet_opts
    return dropdowns
