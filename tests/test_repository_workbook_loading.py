from pathlib import Path
import sys
import importlib
import io
from openpyxl import Workbook


def setup_app(tmp_path, monkeypatch):
    samples = tmp_path / "Samples"
    repo = samples / "repo1Repository"
    repo.mkdir(parents=True)
    wb_path = repo / "CodesetSample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["CODE", "DISPLAY VALUE"])
    wb.save(wb_path)
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    app_module = importlib.import_module("codeset_ui_app.app")
    monkeypatch.setattr(app_module, "SAMPLES_DIR", samples)
    app_module.refresh_repository_cache()
    return app_module, str(repo.relative_to(samples)), wb_path.name


def test_workbooks_endpoint(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()
    from urllib.parse import quote
    resp = client.get(f"/workbooks/{quote(repo, safe='')}")
    assert resp.status_code == 200
    assert fname in resp.get_json()


def test_load_workbook_via_repo(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()
    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200
    assert bytes(f"<option value=\"{fname}\" selected>", "utf-8") in resp.data


def test_repository_list(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()
    resp = client.get("/")
    assert resp.status_code == 200
    assert f'<option value="{repo}"' in resp.get_data(as_text=True)


def test_import_updates_overwrites_file(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    # load existing workbook via repo selection
    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    from openpyxl import Workbook, load_workbook as xl_load

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["CODE", "DISPLAY VALUE"])
    ws.append(["A", "Alpha"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/import",
        data={"workbook": (buf, fname)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200

    # ensure file on disk was replaced
    wb2 = xl_load(tmp_path / "Samples" / repo / fname)
    ws2 = wb2["Sheet1"]
    assert ws2.max_row == 2
    assert ws2["A2"].value == "A"

    # ensure in-memory data updated
    resp = client.get("/sheet/Sheet1")
    assert resp.get_json()[0]["CODE"] == "A"


def test_import_stage_and_confirm(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    from openpyxl import Workbook, load_workbook as xl_load

    staged_wb = Workbook()
    staged_ws = staged_wb.active
    staged_ws.title = "Sheet1"
    staged_ws.append(["CODE", "DISPLAY VALUE"])
    staged_ws.append(["A", "Alpha"])
    staged_ws.append(["B", "Beta"])
    buf = io.BytesIO()
    staged_wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/import",
        data={"stage_only": "1", "workbook": (buf, fname)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["status"] == "pending_changes"
    assert payload["diff"]["summary"]["added"] == 2

    on_disk = xl_load(tmp_path / "Samples" / repo / fname)
    assert on_disk["Sheet1"].max_row == 1

    confirm_resp = client.post("/import/confirm", json={"action": "apply_all"})
    assert confirm_resp.status_code == 200
    assert confirm_resp.get_json()["status"] == "applied"

    updated = xl_load(tmp_path / "Samples" / repo / fname)
    ws = updated["Sheet1"]
    assert ws.max_row == 3
    assert ws["A2"].value == "A"
    assert ws["A3"].value == "B"


def test_import_stage_cancel(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    from openpyxl import Workbook, load_workbook as xl_load

    staged_wb = Workbook()
    staged_ws = staged_wb.active
    staged_ws.title = "Sheet1"
    staged_ws.append(["CODE", "DISPLAY VALUE"])
    staged_ws.append(["A", "Alpha"])
    buf = io.BytesIO()
    staged_wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/import",
        data={"stage_only": "1", "workbook": (buf, fname)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["status"].startswith("pending")

    cancel_resp = client.post("/import/confirm", json={"action": "cancel"})
    assert cancel_resp.status_code == 200
    assert cancel_resp.get_json()["status"] == "cancelled"

    on_disk = xl_load(tmp_path / "Samples" / repo / fname)
    assert on_disk["Sheet1"].max_row == 1


def test_import_stage_sheet_data_includes_added_rows(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    from openpyxl import Workbook

    staged_wb = Workbook()
    staged_ws = staged_wb.active
    staged_ws.title = "Sheet1"
    staged_ws.append(["CODE", "DISPLAY VALUE"])
    staged_ws.append(["A", "Alpha"])
    staged_ws.append(["B", "Beta"])
    buf = io.BytesIO()
    staged_wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/import",
        data={"stage_only": "1", "workbook": (buf, fname)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200

    sheet_resp = client.get("/sheet/Sheet1")
    assert sheet_resp.status_code == 200
    rows = sheet_resp.get_json()
    assert len(rows) == 2
    assert rows[0]["CODE_COMPARE"] == "A"
    assert rows[1]["CODE_COMPARE"] == "B"
    assert rows[1]["CODE"] == ""


def test_upload_form_loads_workbook(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    from openpyxl import Workbook

    staged_wb = Workbook()
    staged_ws = staged_wb.active
    staged_ws.title = "Sheet1"
    staged_ws.append(["CODE", "DISPLAY VALUE"])
    staged_ws.append(["A", "Alpha"])
    staged_ws.append(["B", "Beta"])
    buf = io.BytesIO()
    staged_wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/",
        data={"workbook": (buf, fname)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    html = resp.get_data(as_text=True)
    assert "Pending Workbook Import" not in html
    assert app_module.pending_import_active is False
    assert app_module.workbook_path and app_module.workbook_path.exists()

    sheet_resp = client.get("/sheet/Sheet1")
    assert sheet_resp.status_code == 200
    data = sheet_resp.get_json()
    assert data and data[0]["CODE"] == "A"

def test_export_overwrites_original_file(tmp_path, monkeypatch):
    app_module, repo, fname = setup_app(tmp_path, monkeypatch)
    client = app_module.app.test_client()

    # load workbook and then export with new row
    resp = client.post("/", data={"repo": repo, "workbook_name": fname})
    assert resp.status_code == 200

    payload = {"Sheet1": [{"CODE": "B", "DISPLAY VALUE": "Beta"}]}
    resp = client.post("/export", json=payload)
    assert resp.status_code == 200
    assert resp.get_json()["status"] == "ok"
    assert resp.headers.get("Content-Disposition") is None

    from openpyxl import load_workbook as xl_load

    wb2 = xl_load(tmp_path / "Samples" / repo / fname)
    ws2 = wb2["Sheet1"]
    assert ws2.max_row == 2
    assert ws2["A2"].value == "B"
