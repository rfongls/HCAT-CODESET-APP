"""Tests for sheet protection during export."""

import pandas as pd
from openpyxl import Workbook, load_workbook

from codeset_ui_app.utils.export_excel import export_workbook


def test_export_workbook_leaves_sheet_unprotected_when_false(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["CODE"])

    df = pd.DataFrame([{"CODE": "A"}])
    out = tmp_path / "out.xlsx"

    export_workbook(wb, {"Sheet1": df}, out, False)

    saved = load_workbook(out)
    assert saved["Sheet1"].protection.sheet is False


def test_export_workbook_protects_all_sheets(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["CODE"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["CODE"])

    data = {
        "Sheet1": pd.DataFrame([{"CODE": "A"}]),
        "Sheet2": pd.DataFrame([{"CODE": "B"}]),
    }
    out = tmp_path / "out.xlsx"

    export_workbook(wb, data, out, True)

    saved = load_workbook(out)
    assert saved["Sheet1"].protection.sheet is True
    assert saved["Sheet2"].protection.sheet is True

